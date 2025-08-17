# main.py
from fastapi import FastAPI, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, FileResponse, RedirectResponse, Response
from fastapi.templating import Jinja2Templates
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import tempfile
import os
import logging

from sqlalchemy.orm import Session
import models, schemas
from database import engine, Base, get_db

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

Base.metadata.create_all(bind=engine)

app = FastAPI(title="FastAPI CRUD with Excel Export", version="1.0.0")

# templates folder setup
templates = Jinja2Templates(directory="templates")

# Static files mount removed since no external CSS/JS files are used

# ✅ HTML page
@app.get("/", response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(get_db)):
    items = db.query(models.Item).all()
    return templates.TemplateResponse("index.html", {"request": request, "items": items})

@app.get("/health")
def health_check():
    """Health check endpoint"""
    return {"status": "healthy", "message": "FastAPI CRUD application is running"}

@app.get("/favicon.ico")
def get_favicon():
    """Return an empty favicon to prevent 404 errors"""
    return Response(status_code=204)

@app.get("/robots.txt")
def get_robots():
    """Return robots.txt to prevent 404 errors"""
    return Response(content="User-agent: *\nDisallow: /", media_type="text/plain")

# ✅ REST APIs (same as before)...
@app.post("/items/", response_model=schemas.ItemResponse)
def create_item(item: schemas.ItemCreate, db: Session = Depends(get_db)):
    db_item = models.Item(name=item.name, description=item.description)
    db.add(db_item)
    db.commit()
    db.refresh(db_item)
    return db_item

@app.get("/items/", response_model=list[schemas.ItemResponse])
def read_items(db: Session = Depends(get_db)):
    return db.query(models.Item).all()

@app.put("/items/{item_id}", response_model=schemas.ItemResponse)
def update_item(item_id: int, item: schemas.ItemCreate, db: Session = Depends(get_db)):
    db_item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")
    
    db_item.name = item.name
    db_item.description = item.description
    db.commit()
    db.refresh(db_item)
    return db_item

@app.delete("/items/group")
def delete_multiple_items(request: schemas.GroupDeleteRequest, db: Session = Depends(get_db)):
    """Delete multiple items by their IDs"""
    logging.info(f"Group delete request received for IDs: {request.item_ids}")
    
    if not request.item_ids:
        logging.warning("No item IDs provided in request")
        raise HTTPException(status_code=400, detail="No item IDs provided")
    
    # Validate that all IDs are positive integers
    if any(id <= 0 for id in request.item_ids):
        logging.warning(f"Invalid IDs found: {[id for id in request.item_ids if id <= 0]}")
        raise HTTPException(status_code=400, detail="All item IDs must be positive integers")
    
    # Find and delete all items with the specified IDs
    items_to_delete = db.query(models.Item).filter(models.Item.id.in_(request.item_ids)).all()
    logging.info(f"Found {len(items_to_delete)} items to delete")
    
    if not items_to_delete:
        logging.warning(f"No items found for IDs: {request.item_ids}")
        raise HTTPException(status_code=404, detail="No items found with the provided IDs")
    
    # Check if some IDs were not found
    found_ids = [item.id for item in items_to_delete]
    not_found_ids = [id for id in request.item_ids if id not in found_ids]
    
    if not_found_ids:
        logging.warning(f"Some IDs not found: {not_found_ids}")
        raise HTTPException(
            status_code=404, 
            detail=f"Items with IDs {not_found_ids} not found"
        )
    
    # Delete all found items
    for item in items_to_delete:
        db.delete(item)
        logging.info(f"Deleted item with ID: {item.id}")
    
    db.commit()
    logging.info(f"Successfully deleted {len(items_to_delete)} items")
    return {"message": f"Successfully deleted {len(items_to_delete)} items"}

@app.delete("/items/{item_id}")
def delete_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    db.delete(item)
    db.commit()
    return {"message": "Item deleted successfully"}

@app.get("/download-excel-template")
def download_excel_template():
    """Download Excel template with field names and instructions"""
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Template"
    
    # Add instructions row
    ws.cell(row=1, column=1, value="📝 INSTRUCTIONS:")
    ws.cell(row=1, column=2, value="• Leave ID empty for new items")
    ws.cell(row=1, column=3, value="• Use existing ID to edit that item")
    ws.cell(row=1, column=4, value="• Invalid IDs fall back to name matching")
    
    # Style instructions row
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Add headers with styling
    headers = ["ID", "Name", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add sample data rows
    sample_data = [
        ["", "New Item", "This will create a new item"],
        ["1", "Existing Item", "This will update item with ID=1"],
        ["", "Another New Item", "This will create another new item"]
    ]
    
    for row_idx, (excel_id, name, desc) in enumerate(sample_data, 4):
        ws.cell(row=row_idx, column=1, value=excel_id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=desc)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    return FileResponse(
        path=tmp_path,
        filename="items_template.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/download-excel-data")
def download_excel_data(db: Session = Depends(get_db)):
    """Download all items data as Excel file with editing instructions"""
    # Get all items from database
    items = db.query(models.Item).all()
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Items Data"
    
    # Add instructions row
    ws.cell(row=1, column=1, value="📝 EDITING INSTRUCTIONS:")
    ws.cell(row=1, column=2, value="• Edit Name/Description to update existing items")
    ws.cell(row=1, column=3, value="• Leave ID empty for new items")
    ws.cell(row=1, column=4, value="• Upload back to apply changes")
    
    # Style instructions row
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Add headers with styling
    headers = ["ID", "Name", "Description"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row, item in enumerate(items, 4):
        ws.cell(row=row, column=1, value=item.id)
        ws.cell(row=row, column=2, value=item.name)
        ws.cell(row=row, column=3, value=item.description)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name
    
    return FileResponse(
        path=tmp_path,
        filename="items_data.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.post("/upload-excel")
async def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Upload Excel file and import data - ID-based editing with fallback to name matching"""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")
    
    try:
        # Read the uploaded file
        contents = await file.read()
        
        # Create a temporary file to work with openpyxl
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(contents)
            tmp_path = tmp.name
        
        # Open the workbook
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active
        
        # Determine starting row based on file structure
        # Check if first row contains instructions
        first_row_value = ws.cell(row=1, column=1).value
        if first_row_value and "INSTRUCTIONS" in str(first_row_value):
            start_row = 4  # Skip instruction row + header row
        else:
            start_row = 2  # Skip only header row
        
        # Process rows
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        
        for row_num, row in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
            # Skip empty rows
            if not any(row):
                continue
                
            # Extract values from row
            excel_id = row[0] if row[0] is not None else None
            name = row[1].strip() if row[1] else ""
            description = row[2].strip() if row[2] else ""
            
            # Validate required fields
            if not name or not description:
                logging.warning(f"Row {row_num}: Missing name or description, skipping")
                skipped_count += 1
                continue
            
            # ID-based editing logic
            if excel_id is not None and excel_id != "":
                try:
                    # Try to find item by ID first
                    existing_item = db.query(models.Item).filter(models.Item.id == int(excel_id)).first()
                    
                    if existing_item:
                        # Update existing item by ID
                        existing_item.name = name
                        existing_item.description = description
                        updated_count += 1
                        logging.info(f"Updated item by ID {excel_id}: {name}")
                    else:
                        # ID not found, create new item with specified ID (if possible)
                        try:
                            # Try to create item with the specified ID
                            new_item = models.Item(id=int(excel_id), name=name, description=description)
                            db.add(new_item)
                            imported_count += 1
                            logging.info(f"Created new item with ID {excel_id}: {name}")
                        except Exception as id_error:
                            # If ID assignment fails, create with auto-generated ID
                            logging.warning(f"Could not assign ID {excel_id}, creating with auto-generated ID: {str(id_error)}")
                            new_item = models.Item(name=name, description=description)
                            db.add(new_item)
                            imported_count += 1
                            logging.info(f"Created new item with auto-generated ID: {name}")
                except (ValueError, TypeError):
                    # Invalid ID format, fall back to name matching
                    logging.warning(f"Invalid ID format '{excel_id}' in row {row_num}, falling back to name matching")
                    existing_item = db.query(models.Item).filter(models.Item.name == name).first()
                    
                    if existing_item:
                        # Update existing item by name
                        existing_item.description = description
                        updated_count += 1
                        logging.info(f"Updated existing item by name: {name}")
                    else:
                        # Create new item
                        new_item = models.Item(name=name, description=description)
                        db.add(new_item)
                        imported_count += 1
                        logging.info(f"Created new item: {name}")
            else:
                # No ID provided, use name matching
                existing_item = db.query(models.Item).filter(models.Item.name == name).first()
                
                if existing_item:
                    # Update existing item by name
                    existing_item.description = description
                    updated_count += 1
                    logging.info(f"Updated existing item by name: {name}")
                else:
                    # Create new item
                    new_item = models.Item(name=name, description=description)
                    db.add(new_item)
                    imported_count += 1
                    logging.info(f"Created new item: {name}")
        
        # Commit all changes
        db.commit()
        
        # Clean up temporary file
        os.unlink(tmp_path)
        
        # Build detailed message
        message_parts = []
        if imported_count > 0:
            message_parts.append(f"{imported_count} new items created")
        if updated_count > 0:
            message_parts.append(f"{updated_count} existing items updated")
        if skipped_count > 0:
            message_parts.append(f"{skipped_count} rows skipped")
        
        message = f"Successfully processed Excel file: {', '.join(message_parts)}"
        logging.info(message)
        return {"message": message}
        
    except Exception as e:
        # Clean up temporary file in case of error
        if 'tmp_path' in locals():
            os.unlink(tmp_path)
        logging.error(f"Error processing Excel file: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error processing Excel file: {str(e)}")

# Catch-all endpoint for 404 requests
@app.exception_handler(404)
async def not_found_handler(request: Request, exc: HTTPException):
    """Handle 404 errors gracefully"""
    if request.url.path in ["/favicon.ico", "/robots.txt"]:
        return Response(status_code=204)
    return {"detail": "Not found", "path": request.url.path}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
